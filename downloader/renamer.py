import openpyxl
import sys
import traceback

#import xlrd
#from xlrd import open_workbook
import shutil, os
from os import listdir
from os.path import isfile, join
#import pandas as pd


#excelFolder = 'F://SecExcelDownload/xlsx/'
excelFolder = 'F:/SecExcelDownloadXLS/test/'
onlyfiles = [f for f in listdir(excelFolder) if isfile(join(excelFolder, f))]
count = 0
for excelFile in onlyfiles:
    # xlsx file and xls file
    count+=1
    if('.xlsx' in excelFile):
        # Use openpyxl
        fileName = excelFile.replace('.xlsx','')
        # fileName = fileName.replace('.xls','')
        wb = openpyxl.load_workbook(excelFolder+excelFile)
        docInfoSheetName = wb.get_sheet_names()[0]
        docInfoSheet = wb.get_sheet_by_name(docInfoSheetName)
        """
        Document Type: 10-Q/10-K/ 10-K-A
        Document Fiscal Year Focus: 2011
        Document Fiscal Period Focus: Q1/Q2/Q3
        """
        typeTemplate = 'Document Type'.lower()
        yearTemplate = 'Document Fiscal Year Focus'.lower()
        periodTemplate = 'Document Fiscal Period Focus'.lower()

        docType = 'null'
        docYear = 'null'
        docPeriod = 'null'
        for i in range(1,50):
            #print(i, docInfoSheet.cell(row=i, column=1).value)
            try:
                # print(docInfoSheet.cell(row=i, column=1).value.lower())
                # print(docInfoSheet.cell(row=i, column=2).value.lower())
                element = docInfoSheet.cell(row=i, column=1).value.lower()
                if(element == typeTemplate):
                    docType = docInfoSheet.cell(row=i, column=2).value
                elif(element == yearTemplate):
                    docYear = docInfoSheet.cell(row=i, column=2).value
                elif(element == periodTemplate):
                    docPeriod = docInfoSheet.cell(row=i, column=2).value
            except:
                continue

        compName = fileName.split('-')[0]
        cik = fileName.split('-')[1]
        docID = fileName.split('-')[2]
        print(compName+'_'+cik+'_'+docType+'_'+str(docYear)+'_'+docPeriod+'_'+docID)

    else:
        # Use xlrd
        a = 1


    # try Panda
    # try:
    #     xls = pd.ExcelFile(excelFolder+excelFile)
    #
    #     sheetX = xls.parse(0) #2 is the sheet number
    #
    #     var1 = sheetX['A']
    #
    #     print(var1[1])
    # except:
    #     traceback.print_exc()
    # break;



    # # just use xlrd
    # fileName = excelFile.replace('.xlsx','')
    # fileName = fileName.replace('.xls','')
    # wb = xlrd.open_workbook(excelFolder+excelFile, on_demand=True)
    # docInfoSheet = wb.sheet_by_index(0)
    # # wb = openpyxl.load_workbook(excelFolder+excelFile)
    # # docInfoSheetName = wb.get_sheet_names()[0]
    # # docInfoSheet = wb.get_sheet_by_name(docInfoSheetName)
    # for i in range(0,30):
    #     print(i, docInfoSheet.cell(i, 0).value)
    # break
    # Rename
    # shutil.move('C:\\bacon.txt', 'C:\\eggs')




    # if 'Document Type' in value:
    #     docType = value
    # else if 'Document Fiscal Year Focus' in value:
    #     year = value
    # else if 'Document Fiscal Period Focus' in value:
    #     quarter = value
